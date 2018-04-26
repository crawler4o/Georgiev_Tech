import csv
import os
import io
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from django.http import HttpResponse


def xlsx_writer(csv_files):
    wb = Workbook()

    for x in csv_files:
        sheet = wb.create_sheet(title=x)
        raw_file = csv_files[x].read().decode('utf-8', 'replace')
        dialect = csv.Sniffer().sniff(raw_file)
        csv_data = csv.reader(raw_file, dialect)
        row_track = 1
        column_track = 1
        for row_ in csv_data:
            for column_ in row_:
                sheet.cell(row=row_track, column=column_track).value = column_
                column_track +=1
            row_track +=1
            column_track = 1
    return wb
